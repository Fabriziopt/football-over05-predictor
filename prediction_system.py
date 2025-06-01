import pandas as pd
import numpy as np
from datetime import datetime, timedelta
from sklearn.metrics import accuracy_score, precision_score, recall_score, f1_score
import requests
import warnings
warnings.filterwarnings('ignore')

# Para exportar Excel
try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

class HTGoalsAPIClient:
    """Cliente para integração com API HT Goals"""
    
    def __init__(self, api_key):
        self.api_key = api_key
        self.base_url = "https://api.htgoals.com/v1"
        self.headers = {
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json"
        }
    
    def get_historical_data(self, league_id, days=365):
        """Busca dados históricos da liga"""
        endpoint = f"{self.base_url}/matches/historical"
        params = {
            "league_id": league_id,
            "days": days,
            "include_stats": True
        }
        
        try:
            response = requests.get(endpoint, headers=self.headers, params=params)
            response.raise_for_status()
            data = response.json()
            
            # Converter para DataFrame
            if 'matches' in data:
                df = pd.DataFrame(data['matches'])
                df['date'] = pd.to_datetime(df['date'])
                df['over_05'] = df['ht_total_goals'] > 0.5
                return df
            return pd.DataFrame()
            
        except Exception as e:
            print(f"Erro ao buscar dados históricos: {e}")
            return pd.DataFrame()
    
    def get_team_data(self, team_id, position='both', days=90):
        """Busca dados específicos de um time"""
        endpoint = f"{self.base_url}/teams/{team_id}/matches"
        params = {
            "days": days,
            "position": position,
            "include_ht_stats": True
        }
        
        try:
            response = requests.get(endpoint, headers=self.headers, params=params)
            response.raise_for_status()
            data = response.json()
            
            if 'matches' in data:
                df = pd.DataFrame(data['matches'])
                df['date'] = pd.to_datetime(df['date'])
                df['over_05'] = df['ht_total_goals'] > 0.5
                return df
            return pd.DataFrame()
            
        except Exception as e:
            print(f"Erro ao buscar dados do time: {e}")
            return pd.DataFrame()
    
    def make_prediction(self, match_data):
        """Envia predição para API"""
        endpoint = f"{self.base_url}/predictions/create"
        
        try:
            response = requests.post(endpoint, headers=self.headers, json=match_data)
            response.raise_for_status()
            return response.json()
            
        except Exception as e:
            print(f"Erro ao enviar predição: {e}")
            return None

class SmartPredictionSystem:
    """Sistema de Predição Inteligente com Análise Hierárquica"""
    
    def __init__(self, api_client=None):
        self.api_client = api_client
        self.league_base_rates = {}
        self.team_adjustments = {}
        self.backtesting_results = {}
        
        # Parâmetros otimizáveis
        self.outlier_threshold = 1.5
        self.recent_weight = 0.3
        self.consistency_weight = 0.7
        self.min_games_threshold = 3
        
        # NOVO: Parâmetros da API HT Goals - Confiança 70-100%
        self.confidence_range = {'min': 70, 'max': 100}
    
    def detect_outliers(self, goals_series):
        """Detecta outliers usando IQR method com threshold configurável
        
        MELHORADO: Trata casos especiais como [5, 0, 0, 0, 0]
        """
        if len(goals_series) < 3:
            return np.zeros(len(goals_series), dtype=bool)
        
        q75 = np.percentile(goals_series, 75)
        q25 = np.percentile(goals_series, 25)
        iqr = q75 - q25
        
        # NOVO: Se IQR = 0 (muitos valores iguais), usar método alternativo
        if iqr == 0:
            mean = np.mean(goals_series)
            std = np.std(goals_series)
            if std > 0:
                # Marcar como outlier valores > 2 desvios padrão da média
                outliers = np.abs(goals_series - mean) > (2 * std)
            else:
                # Se todos valores iguais, nenhum é outlier
                outliers = np.zeros(len(goals_series), dtype=bool)
        else:
            # Método IQR tradicional
            lower_bound = q25 - self.outlier_threshold * iqr
            upper_bound = q75 + self.outlier_threshold * iqr
            outliers = (goals_series < lower_bound) | (goals_series > upper_bound)
        
        return outliers
    
    def calculate_realistic_team_stats(self, team_matches, position='home'):
        """Calcula estatísticas realistas removendo outliers e ponderando recência"""
        
        if len(team_matches) == 0:
            return {
                'raw_over_rate': 0.5,
                'adjusted_over_rate': 0.5,
                'consistency_score': 0.5,
                'recent_form': 0.5,
                'scoring_frequency': 0.5,
                'avg_when_scores': 0.5,
                'games_count': 0,
                'outliers_detected': 0
            }
        
        # Determinar coluna de gols baseada na posição
        goals_col = f'ht_{position}_goals' if position in ['home', 'away'] else 'ht_total_goals'
        
        if goals_col not in team_matches.columns:
            return {
                'raw_over_rate': 0.5,
                'adjusted_over_rate': 0.5,
                'consistency_score': 0.5,
                'recent_form': 0.5,
                'scoring_frequency': 0.5,
                'avg_when_scores': 0.5,
                'games_count': 0,
                'outliers_detected': 0
            }
        
        goals_series = team_matches[goals_col].fillna(0)
        over_05_series = team_matches['over_05'].fillna(0)
        
        # 1. Taxa Over 0.5 raw
        raw_over_rate = over_05_series.mean()
        
        # 2. Detectar outliers
        outliers = self.detect_outliers(goals_series)
        
        # 3. Remover outliers para cálculo ajustado
        clean_goals = goals_series[~outliers] if outliers.sum() > 0 else goals_series
        clean_over = over_05_series[~outliers] if outliers.sum() > 0 else over_05_series
        
        adjusted_over_rate = clean_over.mean() if len(clean_over) > 0 else raw_over_rate
        
        # 4. Análise de consistência com peso configurável
        consistency_score = (1 / (1 + np.std(clean_over) + 0.1)) * self.consistency_weight
        
        # 5. Forma recente (últimos 5 jogos) com peso configurável
        recent_games = over_05_series.tail(5)
        recent_form = recent_games.mean() if len(recent_games) > 0 else adjusted_over_rate
        
        # 6. Frequência de marcar
        games_with_goals = len(goals_series[goals_series > 0])
        scoring_frequency = games_with_goals / len(goals_series) if len(goals_series) > 0 else 0
        
        # 7. Média quando marca
        goals_when_scores = goals_series[goals_series > 0]
        avg_when_scores = goals_when_scores.mean() if len(goals_when_scores) > 0 else 0
        
        return {
            'raw_over_rate': raw_over_rate,
            'adjusted_over_rate': adjusted_over_rate,
            'consistency_score': consistency_score,
            'recent_form': recent_form,
            'scoring_frequency': scoring_frequency,
            'avg_when_scores': avg_when_scores,
            'games_count': len(team_matches),
            'outliers_detected': outliers.sum()
        }
    
    def analyze_league(self, league_data):
        """
        ETAPA 1: Análise da Liga
        Estabelece a taxa base da liga
        """
        if len(league_data) == 0:
            return {'base_rate': 0.5, 'volatility': 0.5, 'games_count': 0, 'seasonal_trend': 0}
        
        # Taxa base da liga
        base_rate = league_data['over_05'].mean()
        
        # Volatilidade da liga (algumas ligas são mais previsíveis)
        volatility = np.std(league_data['over_05'])
        
        # Padrões sazonais (início vs final de temporada)
        if 'date' in league_data.columns:
            league_data['date'] = pd.to_datetime(league_data['date'], errors='coerce')
            recent_data = league_data.tail(int(len(league_data) * 0.3))  # Últimos 30%
            recent_rate = recent_data['over_05'].mean()
            seasonal_trend = recent_rate - base_rate
        else:
            seasonal_trend = 0
        
        return {
            'base_rate': base_rate,
            'volatility': volatility,
            'seasonal_trend': seasonal_trend,
            'games_count': len(league_data)
        }
    
    def analyze_teams(self, home_team_data, away_team_data):
        """
        ETAPA 2: Análise dos Times
        Calcula ajustes baseados no histórico dos times
        """
        
        # Analisar time da casa
        home_stats = self.calculate_realistic_team_stats(home_team_data, 'home')
        
        # Analisar time de fora  
        away_stats = self.calculate_realistic_team_stats(away_team_data, 'away')
        
        # Calcular ajustes com pesos configuráveis
        home_adjustment = (home_stats['adjusted_over_rate'] - 0.5) * home_stats['consistency_score']
        away_adjustment = (away_stats['adjusted_over_rate'] - 0.5) * away_stats['consistency_score']
        
        # Forma recente tem peso extra configurável
        home_recent_adj = (home_stats['recent_form'] - home_stats['adjusted_over_rate']) * self.recent_weight
        away_recent_adj = (away_stats['recent_form'] - away_stats['adjusted_over_rate']) * self.recent_weight
        
        return {
            'home_stats': home_stats,
            'away_stats': away_stats,
            'home_adjustment': home_adjustment + home_recent_adj,
            'away_adjustment': away_adjustment + away_recent_adj,
            'combined_adjustment': (home_adjustment + away_adjustment + home_recent_adj + away_recent_adj) / 2
        }
    
    def analyze_match_context(self, home_stats, away_stats, league_analysis):
        """
        ETAPA 3: Análise do Contexto do Jogo
        Fatores específicos do confronto
        """
        
        # Head-to-head seria aqui (se disponível)
        h2h_adjustment = 0  # Placeholder
        
        # Motivação/contexto (posição na tabela, etc.)
        motivation_adjustment = 0  # Placeholder
        
        # Estilo de jogo compatibilidade
        # Times ofensivos vs defensivos
        home_attack_style = home_stats['scoring_frequency']
        away_defense_style = 1 - away_stats['scoring_frequency']
        
        style_compatibility = (home_attack_style + away_defense_style) / 2 - 0.5
        
        return {
            'h2h_adjustment': h2h_adjustment,
            'motivation_adjustment': motivation_adjustment,
            'style_compatibility': style_compatibility,
            'final_context_adjustment': (h2h_adjustment + motivation_adjustment + style_compatibility) / 3
        }
    
    def predict_match(self, league_data, home_team_data, away_team_data):
        """
        Sistema Hierárquico Completo de Predição
        """
        
        # ETAPA 1: Análise da Liga
        league_analysis = self.analyze_league(league_data)
        base_probability = league_analysis['base_rate']
        
        # ETAPA 2: Análise dos Times
        team_analysis = self.analyze_teams(home_team_data, away_team_data)
        team_adjustment = team_analysis['combined_adjustment']
        
        # ETAPA 3: Análise do Contexto
        context_analysis = self.analyze_match_context(
            team_analysis['home_stats'], 
            team_analysis['away_stats'], 
            league_analysis
        )
        context_adjustment = context_analysis['final_context_adjustment']
        
        # ETAPA 4: Cálculo Final
        # Base da liga + ajustes dos times + contexto do jogo
        final_probability = base_probability + team_adjustment + context_adjustment
        
        # Manter entre 0 e 1
        final_probability = max(0.05, min(0.95, final_probability))
        
        # Convertir para porcentagem
        confidence_percentage = final_probability * 100
        
        # NOVO: Garantir que está dentro dos limites da API (70-100%)
        confidence_percentage = max(self.confidence_range['min'], 
                                  min(self.confidence_range['max'], confidence_percentage))
        
        return {
            'final_probability': final_probability,
            'confidence_percentage': confidence_percentage,
            'breakdown': {
                'league_base': base_probability * 100,
                'team_adjustment': team_adjustment * 100,
                'context_adjustment': context_adjustment * 100,
                'final': confidence_percentage
            },
            'analysis_details': {
                'league': league_analysis,
                'teams': team_analysis,
                'context': context_analysis
            }
        }
    
    # NOVA FUNCIONALIDADE: Sistema Unificado com Validação
    def unified_prediction_with_validation(self, league_data, home_team_data, away_team_data):
        """
        Sistema Unificado de Predição com Validação Histórica
        
        Processo completo:
        1. Análise da Liga (base rate)
        2. Análise dos Times (ajustes específicos)
        3. Contexto do Jogo (confronto direto)
        4. Comparação com Histórico (validação)
        5. Ajuste Final da Confiança
        """
        
        print("🔄 Iniciando Análise Unificada...")
        
        # ETAPA 1: Predição Base
        base_prediction = self.predict_match(league_data, home_team_data, away_team_data)
        
        # ETAPA 2: Validação com Histórico Similar
        historical_validation = self.validate_with_historical_matches(
            league_data, 
            base_prediction['confidence_percentage'],
            base_prediction['breakdown']
        )
        
        # ETAPA 3: Ajuste de Confiança baseado no Histórico
        adjusted_confidence = self.adjust_confidence_by_history(
            base_prediction['confidence_percentage'],
            historical_validation['accuracy_in_range'],
            historical_validation['matches_count']
        )
        
        # ETAPA 4: Decisão Final
        recommendation = self.make_final_recommendation(
            adjusted_confidence,
            historical_validation
        )
        
        return {
            'final_confidence': adjusted_confidence,
            'base_confidence': base_prediction['confidence_percentage'],
            'historical_accuracy': historical_validation['accuracy_in_range'],
            'similar_matches_analyzed': historical_validation['matches_count'],
            'recommendation': recommendation,
            'breakdown': base_prediction['breakdown'],
            'validation_details': historical_validation,
            'analysis_summary': {
                'league_tendency': f"{base_prediction['breakdown']['league_base']:.1f}%",
                'teams_adjustment': f"{base_prediction['breakdown']['team_adjustment']:.1f}%",
                'historical_performance': f"{historical_validation['accuracy_in_range']:.1f}%",
                'confidence_level': self.get_confidence_level(adjusted_confidence)
            }
        }
    
    def validate_with_historical_matches(self, league_data, predicted_confidence, breakdown):
        """
        Valida predição comparando com jogos históricos similares
        """
        # Filtrar jogos com características similares
        confidence_range = (predicted_confidence - 5, predicted_confidence + 5)
        
        similar_matches = []
        
        for idx, match in league_data.iterrows():
            # Simular predição para jogos passados
            try:
                # Pegar dados históricos até aquele jogo
                historical_until_match = league_data[league_data['date'] < match['date']]
                
                if len(historical_until_match) < 20:
                    continue
                
                # Dados dos times
                home_data = historical_until_match[
                    historical_until_match['home_team_id'] == match['home_team_id']
                ]
                away_data = historical_until_match[
                    historical_until_match['away_team_id'] == match['away_team_id']
                ]
                
                if len(home_data) >= self.min_games_threshold and len(away_data) >= self.min_games_threshold:
                    # Fazer predição retroativa
                    retro_prediction = self.predict_match(historical_until_match, home_data, away_data)
                    
                    # Se confiança similar, adicionar aos similares
                    if confidence_range[0] <= retro_prediction['confidence_percentage'] <= confidence_range[1]:
                        similar_matches.append({
                            'predicted_conf': retro_prediction['confidence_percentage'],
                            'actual_result': match['over_05'],
                            'correct': (retro_prediction['final_probability'] > 0.5) == match['over_05']
                        })
                        
            except:
                continue
        
        # Calcular estatísticas dos jogos similares
        if len(similar_matches) >= 5:  # Mínimo de jogos para validação
            correct_predictions = sum([m['correct'] for m in similar_matches])
            accuracy = (correct_predictions / len(similar_matches)) * 100
            
            return {
                'matches_count': len(similar_matches),
                'accuracy_in_range': accuracy,
                'confidence_range': confidence_range,
                'avg_confidence': np.mean([m['predicted_conf'] for m in similar_matches])
            }
        
        return {
            'matches_count': 0,
            'accuracy_in_range': predicted_confidence,  # Mantém confiança original
            'confidence_range': confidence_range,
            'avg_confidence': predicted_confidence
        }
    
    def adjust_confidence_by_history(self, base_confidence, historical_accuracy, matches_count):
        """
        Ajusta confiança baseado no desempenho histórico
        """
        if matches_count < 5:
            # Poucos dados históricos, manter confiança base
            return base_confidence
        
        # Peso do histórico aumenta com quantidade de jogos
        historical_weight = min(0.4, matches_count / 50)  # Máximo 40% de peso
        
        # Ajuste ponderado
        adjusted = (base_confidence * (1 - historical_weight)) + (historical_accuracy * historical_weight)
        
        # Garantir limites 70-100%
        return max(self.confidence_range['min'], min(self.confidence_range['max'], adjusted))
    
    def make_final_recommendation(self, confidence, validation):
        """
        Faz recomendação final baseada em todos os fatores
        """
        if validation['matches_count'] < 5:
            if confidence >= 85:
                return "BET_CAUTIOUS"  # Apostar com cautela (poucos dados históricos)
            else:
                return "ANALYZE_MORE"  # Precisa mais análise
        
        # Com dados históricos suficientes
        if confidence >= 85 and validation['accuracy_in_range'] >= 75:
            return "STRONG_BET"  # Aposta forte
        elif confidence >= 75 and validation['accuracy_in_range'] >= 70:
            return "MODERATE_BET"  # Aposta moderada
        elif confidence >= 70:
            return "WEAK_BET"  # Aposta fraca
        else:
            return "SKIP"  # Pular esta aposta
    
    def get_confidence_level(self, confidence):
        """
        Retorna nível de confiança em texto
        """
        if confidence >= 90:
            return "MUITO ALTA"
        elif confidence >= 80:
            return "ALTA"
        elif confidence >= 70:
            return "MODERADA"
        else:
            return "BAIXA"
    
    # NOVA FUNCIONALIDADE: Comparação com Baseline da Liga
    def compare_model_vs_league_baseline(self, backtest_results, league_data):
        """
        Compara performance do modelo vs taxa base da liga
        Isso mostra se o modelo está agregando valor real
        """
        # Taxa base da liga (% de jogos Over 0.5 HT)
        league_baseline = league_data['over_05'].mean() * 100
        
        # Performance do modelo
        model_accuracy = backtest_results.get('overall_accuracy', 0)
        
        # Calcular lift (melhoria sobre baseline)
        if league_baseline > 0:
            lift = ((model_accuracy - league_baseline) / league_baseline) * 100
        else:
            lift = 0
        
        # Análise por tipo de aposta
        performance_vs_baseline = {}
        
        for rec_type, stats in backtest_results['by_recommendation'].items():
            if stats['count'] > 0:
                # Taxa de acerto do modelo para este tipo
                model_rate = stats['accuracy']
                
                # Comparar com baseline
                improvement = model_rate - league_baseline
                
                performance_vs_baseline[rec_type] = {
                    'model_accuracy': model_rate,
                    'league_baseline': league_baseline,
                    'improvement': improvement,
                    'relative_lift': (improvement / league_baseline * 100) if league_baseline > 0 else 0
                }
        
        return {
            'league_baseline': league_baseline,
            'model_overall_accuracy': model_accuracy,
            'absolute_improvement': model_accuracy - league_baseline,
            'relative_lift': lift,
            'by_recommendation': performance_vs_baseline,
            'is_model_better': model_accuracy > league_baseline
        }
    
    # NOVA FUNCIONALIDADE: Integração com API HT Goals
    def predict_from_api(self, league_id, home_team_id, away_team_id, match_date=None):
        """
        Faz predição completa usando dados da API HT Goals
        """
        if not self.api_client:
            raise ValueError("API client não configurado")
        
        # Buscar dados da API
        league_data = self.api_client.get_historical_data(league_id, days=365)
        home_team_data = self.api_client.get_team_data(home_team_id, position='home', days=90)
        away_team_data = self.api_client.get_team_data(away_team_id, position='away', days=90)
        
        # Verificar dados mínimos
        if len(home_team_data) < self.min_games_threshold:
            return {
                'error': f'Dados insuficientes para time da casa (mínimo: {self.min_games_threshold} jogos)',
                'confidence_percentage': 70
            }
        
        if len(away_team_data) < self.min_games_threshold:
            return {
                'error': f'Dados insuficientes para time visitante (mínimo: {self.min_games_threshold} jogos)',
                'confidence_percentage': 70
            }
        
        # Fazer predição unificada
        prediction = self.unified_prediction_with_validation(league_data, home_team_data, away_team_data)
        
        # Preparar dados para API
        api_payload = {
            'league_id': league_id,
            'home_team_id': home_team_id,
            'away_team_id': away_team_id,
            'match_date': match_date or datetime.now().isoformat(),
            'prediction': {
                'market': 'over_05_ht',
                'confidence': prediction['final_confidence'],
                'recommendation': prediction['recommendation'],
                'breakdown': prediction['breakdown']
            },
            'metadata': {
                'model_version': '2.0',
                'analysis_type': 'unified_hierarchical',
                'data_points': {
                    'league_games': len(league_data),
                    'home_games': len(home_team_data),
                    'away_games': len(away_team_data),
                    'similar_matches': prediction['similar_matches_analyzed']
                }
            }
        }
        
        # Enviar para API
        api_response = self.api_client.make_prediction(api_payload)
        prediction['api_response'] = api_response
        
        return prediction

class BacktestingEngine:
    """Engine para testar performance histórica com 365 dias e divisão train/val/test"""
    
    def __init__(self, prediction_system):
        self.prediction_system = prediction_system
        self.results = []
        self.training_history = []
    
    def split_temporal_data(self, historical_data, train_days=240, val_days=60, test_days=65):
        """
        Divisão temporal dos dados (365 dias total):
        - Train: 240 dias (65.8%) - Para treinar o sistema
        - Validation: 60 dias (16.4%) - Para ajustar parâmetros  
        - Test: 65 dias (17.8%) - Para avaliar performance final
        """
        
        if 'date' not in historical_data.columns:
            raise ValueError("Dados históricos devem ter coluna 'date'")
        
        # Ordenar por data
        historical_data['date'] = pd.to_datetime(historical_data['date'], errors='coerce')
        historical_data = historical_data.sort_values('date').reset_index(drop=True)
        
        # Calcular datas de corte
        max_date = historical_data['date'].max()
        
        test_start = max_date - timedelta(days=test_days)
        val_start = test_start - timedelta(days=val_days)
        train_start = val_start - timedelta(days=train_days)
        
        # Filtrar apenas últimos 365 dias
        cutoff_date = max_date - timedelta(days=365)
        recent_data = historical_data[historical_data['date'] >= cutoff_date].copy()
        
        # Dividir dados
        train_data = recent_data[
            (recent_data['date'] >= train_start) & 
            (recent_data['date'] < val_start)
        ].copy()
        
        val_data = recent_data[
            (recent_data['date'] >= val_start) & 
            (recent_data['date'] < test_start)
        ].copy()
        
        test_data = recent_data[
            recent_data['date'] >= test_start
        ].copy()
        
        return {
            'train': train_data,
            'validation': val_data, 
            'test': test_data,
            'split_info': {
                'train_period': f"{train_start.strftime('%Y-%m-%d')} to {val_start.strftime('%Y-%m-%d')}",
                'val_period': f"{val_start.strftime('%Y-%m-%d')} to {test_start.strftime('%Y-%m-%d')}",
                'test_period': f"{test_start.strftime('%Y-%m-%d')} to {max_date.strftime('%Y-%m-%d')}",
                'train_games': len(train_data),
                'val_games': len(val_data),
                'test_games': len(test_data)
            }
        }
    
    def optimize_parameters(self, train_data, val_data):
        """
        Otimiza parâmetros do sistema usando dados de validação
        """
        
        best_accuracy = 0
        best_params = {
            'outlier_threshold': 1.5,
            'recent_weight': 0.3,
            'consistency_weight': 0.7,
            'min_games_threshold': 3
        }
        
        # Grid search para otimizar parâmetros
        outlier_thresholds = [1.0, 1.5, 2.0]
        recent_weights = [0.2, 0.3, 0.4]
        consistency_weights = [0.6, 0.7, 0.8]
        min_games = [3, 5, 7]
        
        optimization_results = []
        
        print("🔧 Testando combinações de parâmetros...")
        total_combinations = len(outlier_thresholds) * len(recent_weights) * len(consistency_weights) * len(min_games)
        current_combination = 0
        
        for outlier_thresh in outlier_thresholds:
            for recent_w in recent_weights:
                for consist_w in consistency_weights:
                    for min_g in min_games:
                        current_combination += 1
                        
                        if current_combination % 20 == 0:
                            print(f"   Progresso: {current_combination}/{total_combinations}")
                        
                        # Testar estes parâmetros
                        temp_system = SmartPredictionSystem()
                        temp_system.outlier_threshold = outlier_thresh
                        temp_system.recent_weight = recent_w
                        temp_system.consistency_weight = consist_w
                        temp_system.min_games_threshold = min_g
                        
                        # Avaliar no conjunto de validação
                        val_accuracy = self.evaluate_on_dataset(temp_system, train_data, val_data)
                        
                        optimization_results.append({
                            'outlier_threshold': outlier_thresh,
                            'recent_weight': recent_w,
                            'consistency_weight': consist_w,
                            'min_games_threshold': min_g,
                            'val_accuracy': val_accuracy
                        })
                        
                        if val_accuracy > best_accuracy:
                            best_accuracy = val_accuracy
                            best_params = {
                                'outlier_threshold': outlier_thresh,
                                'recent_weight': recent_w,
                                'consistency_weight': consist_w,
                                'min_games_threshold': min_g
                            }
        
        return best_params, best_accuracy, optimization_results
    
    def evaluate_on_dataset(self, system, train_data, eval_data):
        """Avalia sistema em um conjunto de dados"""
        
        predictions = []
        actuals = []
        
        # Agrupar por liga para eficiência
        leagues = eval_data['league_id'].unique()
        
        for league_id in leagues:
            league_train = train_data[train_data['league_id'] == league_id]
            league_eval = eval_data[eval_data['league_id'] == league_id]
            
            if len(league_train) < 20:  # Mínimo para treinamento
                continue
            
            # Para cada jogo de avaliação
            for idx, match in league_eval.iterrows():
                try:
                    # Dados dos times até a data do jogo
                    match_date = match['date']
                    historical_until_match = league_train[league_train['date'] < match_date]
                    
                    if len(historical_until_match) < 10:
                        continue
                    
                    # Dados específicos dos times
                    home_team_data = historical_until_match[
                        historical_until_match['home_team_id'] == match['home_team_id']
                    ]
                    away_team_data = historical_until_match[
                        historical_until_match['away_team_id'] == match['away_team_id']
                    ]
                    
                    if len(home_team_data) < system.min_games_threshold or \
                       len(away_team_data) < system.min_games_threshold:
                        continue
                    
                    # Fazer predição
                    prediction = system.predict_match(
                        historical_until_match, home_team_data, away_team_data
                    )
                    
                    # Comparar com resultado real
                    predicted_class = prediction['final_probability'] > 0.5
                    actual_class = match['over_05'] == 1
                    
                    predictions.append(predicted_class)
                    actuals.append(actual_class)
                    
                except Exception as e:
                    continue
        
        # Calcular acurácia
        if len(predictions) > 0:
            accuracy = accuracy_score(actuals, predictions)
            return accuracy
        
        return 0.0
    
    def run_comprehensive_backtest(self, historical_data):
        """
        Executa backtesting completo com 365 dias:
        1. Split train/val/test
        2. Otimização de parâmetros 
        3. Avaliação final
        4. Análise detalhada
        5. NOVO: Comparação com baseline da liga
        """
        
        print("🚀 Iniciando Backtesting Completo de 365 dias...")
        
        # 1. Dividir dados temporalmente
        print("📊 Dividindo dados em Train/Validation/Test...")
        data_split = self.split_temporal_data(historical_data)
        
        train_data = data_split['train']
        val_data = data_split['validation']
        test_data = data_split['test']
        
        print(f"✅ Train: {len(train_data)} jogos | Val: {len(val_data)} jogos | Test: {len(test_data)} jogos")
        
        # 2. Otimizar parâmetros usando train+validation
        print("🔧 Otimizando parâmetros do sistema...")
        best_params, best_val_acc, optimization_results = self.optimize_parameters(train_data, val_data)
        
        print(f"✅ Melhor acurácia validação: {best_val_acc:.1%}")
        print(f"✅ Melhores parâmetros: {best_params}")
        
        # 3. Aplicar melhores parâmetros ao sistema
        optimized_system = SmartPredictionSystem()
        for param, value in best_params.items():
            setattr(optimized_system, param, value)
        
        # 4. Avaliação final no conjunto de teste
        print("🎯 Avaliação final no conjunto de teste...")
        test_results = self.detailed_evaluation(optimized_system, train_data, val_data, test_data)
        
        # 5. Análise por faixas de confiança
        confidence_analysis = self.analyze_by_confidence_ranges(test_results['detailed_predictions'])
        
        # 6. Análise por liga
        league_analysis = self.analyze_by_league(test_results['detailed_predictions'])
        
        # 7. NOVO: Comparar com baseline da liga
        league_comparison = optimized_system.compare_model_vs_league_baseline(test_results, historical_data)
        
        return {
            'data_split_info': data_split['split_info'],
            'optimization_results': {
                'best_params': best_params,
                'best_val_accuracy': best_val_acc,
                'all_combinations': optimization_results
            },
            'test_results': test_results,
            'confidence_analysis': confidence_analysis,
            'league_analysis': league_analysis,
            'league_comparison': league_comparison,  # NOVO
            'summary': {
                'total_test_predictions': len(test_results['detailed_predictions']),
                'test_accuracy': test_results['accuracy'],
                'test_precision': test_results['precision'],
                'test_recall': test_results['recall'],
                'test_f1': test_results['f1_score'],
                'baseline_improvement': league_comparison['absolute_improvement']  # NOVO
            }
        }
    
    def detailed_evaluation(self, system, train_data, val_data, test_data):
        """Avaliação detalhada com múltiplas métricas"""
        
        # Combinar train+val para treinamento final
        training_data = pd.concat([train_data, val_data]).sort_values('date')
        
        detailed_predictions = []
        predictions = []
        actuals = []
        probabilities = []
        
        # NOVO: Adicionar estatísticas por tipo de recomendação
        by_recommendation = {
            'STRONG_BET': {'count': 0, 'correct': 0, 'accuracy': 0},
            'MODERATE_BET': {'count': 0, 'correct': 0, 'accuracy': 0},
            'WEAK_BET': {'count': 0, 'correct': 0, 'accuracy': 0},
            'SKIP': {'count': 0, 'correct': 0, 'accuracy': 0}
        }
        
        # Agrupar por liga
        leagues = test_data['league_id'].unique()
        
        for league_id in leagues:
            league_training = training_data[training_data['league_id'] == league_id]
            league_test = test_data[test_data['league_id'] == league_id]
            
            if len(league_training) < 30:  # Mínimo para avaliação confiável
                continue
            
            league_name = league_test['league_name'].iloc[0] if 'league_name' in league_test.columns else f"Liga_{league_id}"
            
            # Para cada jogo de teste
            for idx, match in league_test.iterrows():
                try:
                    # Dados históricos até a data do jogo
                    match_date = match['date']
                    historical_data = league_training[league_training['date'] < match_date]
                    
                    # Dados dos times
                    home_team_data = historical_data[
                        historical_data['home_team_id'] == match['home_team_id']
                    ]
                    away_team_data = historical_data[
                        historical_data['away_team_id'] == match['away_team_id']
                    ]
                    
                    if len(home_team_data) < system.min_games_threshold or \
                       len(away_team_data) < system.min_games_threshold:
                        continue
                    
                    # Fazer predição
                    prediction = system.predict_match(historical_data, home_team_data, away_team_data)
                    
                    # Resultado real
                    actual = match['over_05'] == 1
                    predicted = prediction['final_probability'] > 0.5
                    
                    # Salvar resultados detalhados
                    detailed_predictions.append({
                        'date': match['date'],
                        'league': league_name,
                        'home_team': match.get('home_team', 'Unknown'),
                        'away_team': match.get('away_team', 'Unknown'),
                        'predicted_probability': prediction['final_probability'],
                        'confidence_percentage': prediction['confidence_percentage'],
                        'predicted_class': predicted,
                        'actual_class': actual,
                        'correct': predicted == actual,
                        'breakdown': prediction['breakdown'],
                        'ht_home_goals': match.get('ht_home_goals', 0),
                        'ht_away_goals': match.get('ht_away_goals', 0)
                    })
                    
                    predictions.append(predicted)
                    actuals.append(actual)
                    probabilities.append(prediction['final_probability'])
                    
                except Exception as e:
                    continue
        
        # Calcular métricas
        if len(predictions) > 0:
            accuracy = accuracy_score(actuals, predictions)
            precision = precision_score(actuals, predictions, zero_division=0)
            recall = recall_score(actuals, predictions, zero_division=0)
            f1 = f1_score(actuals, predictions, zero_division=0)
            
            # Calcular acurácia por recomendação
            for rec_type, stats in by_recommendation.items():
                if stats['count'] > 0:
                    stats['accuracy'] = (stats['correct'] / stats['count']) * 100
            
            return {
                'accuracy': accuracy,
                'precision': precision,
                'recall': recall,
                'f1_score': f1,
                'total_predictions': len(predictions),
                'detailed_predictions': detailed_predictions,
                'by_recommendation': by_recommendation  # NOVO
            }
        
        return {
            'accuracy': 0,
            'precision': 0,
            'recall': 0,
            'f1_score': 0,
            'total_predictions': 0,
            'detailed_predictions': [],
            'by_recommendation': by_recommendation
        }
    
    def analyze_by_confidence_ranges(self, detailed_predictions):
        """Analisa performance por faixas de confiança - ATUALIZADO para 70-100%"""
        
        if not detailed_predictions:
            return []
        
        df = pd.DataFrame(detailed_predictions)
        
        # NOVO: Faixas de confiança ajustadas para 70-100%
        confidence_ranges = [
            (70, 75), (75, 80), (80, 85), (85, 90), (90, 100)
        ]
        
        analysis = []
        
        for min_conf, max_conf in confidence_ranges:
            mask = (df['confidence_percentage'] >= min_conf) & \
                   (df['confidence_percentage'] < max_conf)
            
            subset = df[mask]
            
            if len(subset) > 0:
                accuracy = subset['correct'].mean()
                count = len(subset)
                avg_confidence = subset['confidence_percentage'].mean()
                
                # Análise adicional
                over_predictions = subset[subset['predicted_class'] == True]
                over_accuracy = over_predictions['correct'].mean() if len(over_predictions) > 0 else 0
                
                analysis.append({
                    'confidence_range': f"{min_conf}-{max_conf}%",
                    'predictions_count': count,
                    'accuracy': accuracy,
                    'avg_confidence': avg_confidence,
                    'over_predictions': len(over_predictions),
                    'over_accuracy': over_accuracy
                })
        
        return analysis
    
    def analyze_by_league(self, detailed_predictions):
        """Analisa performance por liga"""
        
        if not detailed_predictions:
            return []
        
        df = pd.DataFrame(detailed_predictions)
        
        league_analysis = []
        
        for league in df['league'].unique():
            league_data = df[df['league'] == league]
            
            if len(league_data) >= 10:  # Mínimo para análise
                accuracy = league_data['correct'].mean()
                avg_confidence = league_data['confidence_percentage'].mean()
                
                league_analysis.append({
                    'league': league,
                    'predictions_count': len(league_data),
                    'accuracy': accuracy,
                    'avg_confidence': avg_confidence
                })
        
        # Ordenar por acurácia
        league_analysis.sort(key=lambda x: x['accuracy'], reverse=True)
        
        return league_analysis

# NOVA FUNCIONALIDADE: Funções auxiliares para Export e Demonstração
def export_predictions_to_excel(predictions_df, filename="ht_goals_predictions.xlsx"):
    """
    Exporta predições para Excel com formatação profissional
    """
    if not EXCEL_AVAILABLE:
        # Fallback para CSV se openpyxl não estiver instalado
        predictions_df.to_csv(filename.replace('.xlsx', '.csv'), index=False)
        return filename.replace('.xlsx', '.csv')
    
    # Criar Excel writer
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        # Aba principal com predições
        predictions_df.to_excel(writer, sheet_name='Predições', index=False)
        
        # Pegar workbook e worksheet
        workbook = writer.book
        worksheet = writer.sheets['Predições']
        
        # Estilos
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        # Formatar cabeçalhos
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        
        # Ajustar largura das colunas
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Adicionar formatação condicional para confiança
        for row in range(2, len(predictions_df) + 2):
            if 'confidence_percentage' in predictions_df.columns:
                conf_col = predictions_df.columns.get_loc('confidence_percentage') + 1
                conf_cell = worksheet.cell(row=row, column=conf_col)
                
                try:
                    conf_value = float(conf_cell.value)
                    if conf_value >= 85:
                        conf_cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
                    elif conf_value >= 75:
                        conf_cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
                    else:
                        conf_cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                except:
                    pass
        
        # Criar aba de resumo
        summary_data = create_summary_stats(predictions_df)
        summary_df = pd.DataFrame(summary_data)
        summary_df.to_excel(writer, sheet_name='Resumo', index=False)
    
    return filename

def create_summary_stats(predictions_df):
    """Cria estatísticas resumidas das predições"""
    
    summary = []
    
    # Estatísticas gerais
    summary.append({
        'Métrica': 'Total de Predições',
        'Valor': len(predictions_df)
    })
    
    if 'confidence_percentage' in predictions_df.columns:
        summary.append({
            'Métrica': 'Confiança Média',
            'Valor': f"{predictions_df['confidence_percentage'].mean():.1f}%"
        })
    
    # Por faixas de confiança (70-100%)
    confidence_ranges = [(70, 75), (75, 80), (80, 85), (85, 90), (90, 100)]
    
    if 'confidence_percentage' in predictions_df.columns:
        for min_conf, max_conf in confidence_ranges:
            mask = (predictions_df['confidence_percentage'] >= min_conf) & \
                   (predictions_df['confidence_percentage'] < max_conf)
            count = mask.sum()
            
            summary.append({
                'Métrica': f'Predições {min_conf}-{max_conf}%',
                'Valor': count
            })
    
    # Se houver resultados reais
    if 'correct' in predictions_df.columns:
        accuracy = predictions_df['correct'].mean()
        summary.append({
            'Métrica': 'Acurácia Geral',
            'Valor': f"{accuracy:.1%}"
        })
    
    return summary

def demonstrate_outlier_handling():
    """
    Demonstra como o sistema lida com outliers
    Exemplo: Time com [5, 0, 0, 0, 0] gols HT
    """
    
    print("🔍 DEMONSTRAÇÃO: Tratamento de Outliers")
    print("=" * 60)
    
    # Criar sistema
    system = SmartPredictionSystem()
    
    # Dados do exemplo - time da casa
    home_team_data = pd.DataFrame({
        'ht_home_goals': [5, 0, 0, 0, 0],  # Exatamente o exemplo!
        'over_05': [1, 0, 0, 0, 0],
        'date': pd.date_range('2024-01-01', periods=5)
    })
    
    # Analisar time
    stats = system.calculate_realistic_team_stats(home_team_data, 'home')
    
    print("📊 Dados originais do time:")
    print(f"   Gols HT: {home_team_data['ht_home_goals'].tolist()}")
    print(f"   Média com outlier: {home_team_data['ht_home_goals'].mean():.2f} gols")
    
    print("\n🎯 Após detecção de outliers:")
    print(f"   Outliers detectados: {stats['outliers_detected']}")
    print(f"   Taxa Over 0.5 raw: {stats['raw_over_rate']:.1%}")
    print(f"   Taxa Over 0.5 ajustada: {stats['adjusted_over_rate']:.1%}")
    
    # Mostrar o que aconteceu
    goals = home_team_data['ht_home_goals'].values
    outliers = system.detect_outliers(goals)
    clean_goals = goals[~outliers]
    
    print(f"\n✅ Gols após remover outliers: {clean_goals.tolist()}")
    print(f"   Média sem outlier: {clean_goals.mean():.2f} gols")
    print(f"   Frequência de marcar (limpa): {(clean_goals > 0).mean():.1%}")
    
    return stats

# Funções para integração com Streamlit
def create_prediction_breakdown_display(prediction_result):
    """Cria display detalhado da predição"""
    
    breakdown = prediction_result['breakdown']
    details = prediction_result['analysis_details']
    
    return {
        'Liga Base': f"{breakdown['league_base']:.1f}%",
        'Ajuste Times': f"{breakdown['team_adjustment']:+.1f}%",
        'Contexto Jogo': f"{breakdown['context_adjustment']:+.1f}%",
        'Taxa Final': f"{breakdown['final']:.1f}%",
        'Jogos Casa': details['teams']['home_stats']['games_count'],
        'Jogos Fora': details['teams']['away_stats']['games_count'],
        'Consistência Casa': f"{details['teams']['home_stats']['consistency_score']:.2f}",
        'Consistência Fora': f"{details['teams']['away_stats']['consistency_score']:.2f}",
        'Outliers Casa': details['teams']['home_stats']['outliers_detected'],
        'Outliers Fora': details['teams']['away_stats']['outliers_detected']
    }

def run_system_backtest(historical_data, leagues_to_test=None):
    """Executa backtesting para múltiplas ligas"""
    
    prediction_system = SmartPredictionSystem()
    backtesting_engine = BacktestingEngine(prediction_system)
    
    if leagues_to_test is None:
        leagues_to_test = historical_data['league_id'].unique()[:10]  # Top 10 ligas
    
    results = {}
    
    for league_id in leagues_to_test:
        league_data = historical_data[historical_data['league_id'] == league_id]
        
        if len(league_data) < 100:  # Mínimo de jogos
            continue
        
        try:
            backtest_result = backtesting_engine.run_comprehensive_backtest(league_data)
            
            if backtest_result:
                league_name = league_data['league_name'].iloc[0] if 'league_name' in league_data.columns else f"Liga_{league_id}"
                results[league_name] = backtest_result
                
        except Exception as e:
            continue
    
    return results

# Exemplo de uso
def example_usage():
    """Exemplo de como usar o sistema"""
    
    # Criar sistema
    prediction_system = SmartPredictionSystem()
    
    # Dados fictícios para exemplo
    league_data = pd.DataFrame({
        'over_05': [1, 0, 1, 1, 0, 1, 0, 1, 1, 0] * 10,
        'date': pd.date_range('2024-01-01', periods=100)
    })
    
    home_team_data = pd.DataFrame({
        'ht_home_goals': [5, 0, 0, 0, 0],  # Seu exemplo!
        'over_05': [1, 0, 0, 0, 0],
        'date': pd.date_range('2024-01-01', periods=5)
    })
    
    away_team_data = pd.DataFrame({
        'ht_away_goals': [0, 1, 0, 1, 2],
        'over_05': [0, 1, 0, 1, 1],
        'date': pd.date_range('2024-01-01', periods=5)
    })
    
    # Fazer predição
    result = prediction_system.predict_match(league_data, home_team_data, away_team_data)
    
    print("Resultado da Predição:")
    print(f"Taxa Final: {result['confidence_percentage']:.1f}%")
    print(f"Breakdown: {result['breakdown']}")
    
    return result

# Função principal para execução completa
def main_backtest_analysis(historical_data):
    """Função principal para executar análise completa"""
    
    print("🚀 Sistema de Backtesting Inteligente - 365 dias")
    print("=" * 60)
    
    # Criar sistema
    prediction_system = SmartPredictionSystem()
    backtesting_engine = BacktestingEngine(prediction_system)
    
    # Executar backtesting completo
    results = backtesting_engine.run_comprehensive_backtest(historical_data)
    
    # Exibir resultados
    print("\n📊 RESULTADOS FINAIS:")
    print("=" * 60)
    
    summary = results['summary']
    print(f"Total Previsões: {summary['total_test_predictions']}")
    print(f"Acurácia: {summary['test_accuracy']:.1%}")
    print(f"Precisão: {summary['test_precision']:.1%}")
    print(f"Recall: {summary['test_recall']:.1%}")
    print(f"F1-Score: {summary['test_f1']:.1%}")
    
    # NOVO: Comparação com baseline
    if 'baseline_improvement' in summary:
        print(f"Melhoria sobre baseline: {summary['baseline_improvement']:+.1f}%")
    
    # NOVO: Comparação detalhada com liga
    if 'league_comparison' in results:
        comparison = results['league_comparison']
        print(f"\n🎯 COMPARAÇÃO MODELO vs LIGA:")
        print(f"Taxa Base da Liga: {comparison['league_baseline']:.1f}%")
        print(f"Taxa do Modelo: {comparison['model_overall_accuracy']:.1f}%")
        print(f"Lift Relativo: {comparison['relative_lift']:+.1f}%")
    
    print("\n🎯 ANÁLISE POR CONFIANÇA (70-100%):")
    for conf_analysis in results['confidence_analysis']:
        print(f"{conf_analysis['confidence_range']}: "
              f"{conf_analysis['predictions_count']} jogos → "
              f"{conf_analysis['accuracy']:.1%} acerto")
    
    print("\n🏆 TOP 5 LIGAS:")
    for league_analysis in results['league_analysis'][:5]:
        print(f"{league_analysis['league']}: "
              f"{league_analysis['predictions_count']} jogos → "
              f"{league_analysis['accuracy']:.1%} acerto")
    
    return results

# NOVO: Exemplo de uso com API
def example_usage_with_api():
    """Exemplo de uso com integração API HT Goals"""
    
    # Configurar API client
    api_client = HTGoalsAPIClient(api_key="YOUR_API_KEY_HERE")
    
    # Criar sistema de predição
    prediction_system = SmartPredictionSystem(api_client=api_client)
    
    # Fazer predição para um jogo específico
    result = prediction_system.predict_from_api(
        league_id=1,  # Premier League
        home_team_id=10,  # Manchester United
        away_team_id=20,  # Liverpool
        match_date="2025-06-02T15:00:00"
    )
    
    if 'error' not in result:
        print("🎯 Resultado da Predição:")
        print(f"Confiança Final: {result['final_confidence']:.1f}%")
        print(f"Recomendação: {result['recommendation']}")
        print(f"Performance Histórica: {result['historical_accuracy']:.1f}%")
        print("\n📊 Breakdown:")
        for key, value in result['breakdown'].items():
            print(f"  {key}: {value:.1f}%")
    else:
        print(f"❌ Erro: {result['error']}")
    
    return result

if __name__ == "__main__":
    print("Sistema de Predição Inteligente com Backtesting 365 dias")
    print("Versão 2.0 - Com API HT Goals, Confiança 70-100%, Comparação com Liga")
    print("Para usar, chame main_backtest_analysis(seu_dataframe)")
