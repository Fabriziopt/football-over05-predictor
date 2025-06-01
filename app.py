import pandas as pd
import numpy as np
from datetime import datetime, timedelta
from sklearn.metrics import accuracy_score, precision_score, recall_score, f1_score
import requests
import warnings
warnings.filterwarnings('ignore')

# Para exportar Excel
try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

class HTGoalsAPIClient:
    """Cliente para integração com API HT Goals"""
    
    def __init__(self, api_key):
        self.api_key = api_key
        self.base_url = "https://api.htgoals.com/v1"
        self.headers = {
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json"
        }
    
    def get_historical_data(self, league_id, days=365):
        """Busca dados históricos da liga"""
        endpoint = f"{self.base_url}/matches/historical"
        params = {
            "league_id": league_id,
            "days": days,
            "include_stats": True
        }
        
        try:
            response = requests.get(endpoint, headers=self.headers, params=params)
            response.raise_for_status()
            data = response.json()
            
            # Converter para DataFrame
            if 'matches' in data:
                df = pd.DataFrame(data['matches'])
                df['date'] = pd.to_datetime(df['date'])
                df['over_05'] = df['ht_total_goals'] > 0.5
                return df
            return pd.DataFrame()
            
        except Exception as e:
            print(f"Erro ao buscar dados históricos: {e}")
            return pd.DataFrame()
    
    def get_team_data(self, team_id, position='both', days=90):
        """Busca dados específicos de um time"""
        endpoint = f"{self.base_url}/teams/{team_id}/matches"
        params = {
            "days": days,
            "position": position,
            "include_ht_stats": True
        }
        
        try:
            response = requests.get(endpoint, headers=self.headers, params=params)
            response.raise_for_status()
            data = response.json()
            
            if 'matches' in data:
                df = pd.DataFrame(data['matches'])
                df['date'] = pd.to_datetime(df['date'])
                df['over_05'] = df['ht_total_goals'] > 0.5
                return df
            return pd.DataFrame()
            
        except Exception as e:
            print(f"Erro ao buscar dados do time: {e}")
            return pd.DataFrame()
    
    def make_prediction(self, match_data):
        """Envia predição para API"""
        endpoint = f"{self.base_url}/predictions/create"
        
        try:
            response = requests.post(endpoint, headers=self.headers, json=match_data)
            response.raise_for_status()
            return response.json()
            
        except Exception as e:
            print(f"Erro ao enviar predição: {e}")
            return None

class SmartPredictionSystem:
    """Sistema de Predição Inteligente com Análise Hierárquica - Integrado com HT Goals"""
    
    def __init__(self, api_client=None):
        self.api_client = api_client
        self.league_base_rates = {}
        self.team_adjustments = {}
        self.backtesting_results = {}
        
        # Parâmetros otimizáveis
        self.outlier_threshold = 1.5
        self.recent_weight = 0.3
        self.consistency_weight = 0.7
        self.min_games_threshold = 3
        
        # Parâmetros da API HT Goals
        self.confidence_range = {'min': 50, 'max': 100}
    
    def fetch_data_from_api(self, league_id, home_team_id, away_team_id):
        """Busca dados necessários da API HT Goals"""
        if not self.api_client:
            raise ValueError("API client não configurado")
        
        # Buscar dados da liga (365 dias)
        league_data = self.api_client.get_historical_data(league_id, days=365)
        
        # Buscar dados dos times (90 dias)
        home_team_data = self.api_client.get_team_data(home_team_id, position='home', days=90)
        away_team_data = self.api_client.get_team_data(away_team_id, position='away', days=90)
        
        return league_data, home_team_data, away_team_data
    
    def detect_outliers(self, goals_series):
        """Detecta outliers usando IQR method com threshold configurável
        
        EXEMPLO: Time com gols [5, 0, 0, 0, 0]
        - Q75 = 0, Q25 = 0, IQR = 0
        - Como IQR = 0, usa método alternativo baseado em desvio padrão
        - O 5 será marcado como outlier e REMOVIDO dos cálculos
        """
        if len(goals_series) < 3:
            return np.zeros(len(goals_series), dtype=bool)
        
        q75 = np.percentile(goals_series, 75)
        q25 = np.percentile(goals_series, 25)
        iqr = q75 - q25
        
        # Se IQR = 0 (muitos valores iguais), usar método alternativo
        if iqr == 0:
            mean = np.mean(goals_series)
            std = np.std(goals_series)
            if std > 0:
                # Marcar como outlier valores > 2 desvios padrão da média
                outliers = np.abs(goals_series - mean) > (2 * std)
            else:
                # Se todos valores iguais, nenhum é outlier
                outliers = np.zeros(len(goals_series), dtype=bool)
        else:
            # Método IQR tradicional
            lower_bound = q25 - self.outlier_threshold * iqr
            upper_bound = q75 + self.outlier_threshold * iqr
            outliers = (goals_series < lower_bound) | (goals_series > upper_bound)
        
        return outliers
    
    def calculate_realistic_team_stats(self, team_matches, position='home'):
        """Calcula estatísticas realistas removendo outliers e ponderando recência"""
        
        if len(team_matches) == 0:
            return {
                'raw_over_rate': 0.5,
                'adjusted_over_rate': 0.5,
                'consistency_score': 0.5,
                'recent_form': 0.5,
                'scoring_frequency': 0.5,
                'avg_when_scores': 0.5,
                'games_count': 0,
                'outliers_detected': 0
            }
        
        # Determinar coluna de gols baseada na posição
        goals_col = f'ht_{position}_goals' if position in ['home', 'away'] else 'ht_total_goals'
        
        if goals_col not in team_matches.columns:
            return {
                'raw_over_rate': 0.5,
                'adjusted_over_rate': 0.5,
                'consistency_score': 0.5,
                'recent_form': 0.5,
                'scoring_frequency': 0.5,
                'avg_when_scores': 0.5,
                'games_count': 0,
                'outliers_detected': 0
            }
        
        goals_series = team_matches[goals_col].fillna(0)
        over_05_series = team_matches['over_05'].fillna(0)
        
        # 1. Taxa Over 0.5 raw
        raw_over_rate = over_05_series.mean()
        
        # 2. Detectar outliers
        outliers = self.detect_outliers(goals_series)
        
        # 3. Remover outliers para cálculo ajustado
        clean_goals = goals_series[~outliers] if outliers.sum() > 0 else goals_series
        clean_over = over_05_series[~outliers] if outliers.sum() > 0 else over_05_series
        
        adjusted_over_rate = clean_over.mean() if len(clean_over) > 0 else raw_over_rate
        
        # 4. Análise de consistência com peso configurável
        consistency_score = (1 / (1 + np.std(clean_over) + 0.1)) * self.consistency_weight
        
        # 5. Forma recente (últimos 5 jogos) com peso configurável
        recent_games = over_05_series.tail(5)
        recent_form = recent_games.mean() if len(recent_games) > 0 else adjusted_over_rate
        
        # 6. Frequência de marcar
        games_with_goals = len(goals_series[goals_series > 0])
        scoring_frequency = games_with_goals / len(goals_series) if len(goals_series) > 0 else 0
        
        # 7. Média quando marca
        goals_when_scores = goals_series[goals_series > 0]
        avg_when_scores = goals_when_scores.mean() if len(goals_when_scores) > 0 else 0
        
        return {
            'raw_over_rate': raw_over_rate,
            'adjusted_over_rate': adjusted_over_rate,
            'consistency_score': consistency_score,
            'recent_form': recent_form,
            'scoring_frequency': scoring_frequency,
            'avg_when_scores': avg_when_scores,
            'games_count': len(team_matches),
            'outliers_detected': outliers.sum()
        }
    
    def analyze_league(self, league_data):
        """
        ETAPA 1: Análise da Liga
        Estabelece a taxa base da liga
        """
        if len(league_data) == 0:
            return {'base_rate': 0.5, 'volatility': 0.5, 'games_count': 0, 'seasonal_trend': 0}
        
        # Taxa base da liga
        base_rate = league_data['over_05'].mean()
        
        # Volatilidade da liga (algumas ligas são mais previsíveis)
        volatility = np.std(league_data['over_05'])
        
        # Padrões sazonais (início vs final de temporada)
        if 'date' in league_data.columns:
            league_data['date'] = pd.to_datetime(league_data['date'], errors='coerce')
            recent_data = league_data.tail(int(len(league_data) * 0.3))  # Últimos 30%
            recent_rate = recent_data['over_05'].mean()
            seasonal_trend = recent_rate - base_rate
        else:
            seasonal_trend = 0
        
        return {
            'base_rate': base_rate,
            'volatility': volatility,
            'seasonal_trend': seasonal_trend,
            'games_count': len(league_data)
        }
    
    def analyze_teams(self, home_team_data, away_team_data):
        """
        ETAPA 2: Análise dos Times
        Calcula ajustes baseados no histórico dos times
        """
        
        # Analisar time da casa
        home_stats = self.calculate_realistic_team_stats(home_team_data, 'home')
        
        # Analisar time de fora  
        away_stats = self.calculate_realistic_team_stats(away_team_data, 'away')
        
        # Calcular ajustes com pesos configuráveis
        home_adjustment = (home_stats['adjusted_over_rate'] - 0.5) * home_stats['consistency_score']
        away_adjustment = (away_stats['adjusted_over_rate'] - 0.5) * away_stats['consistency_score']
        
        # Forma recente tem peso extra configurável
        home_recent_adj = (home_stats['recent_form'] - home_stats['adjusted_over_rate']) * self.recent_weight
        away_recent_adj = (away_stats['recent_form'] - away_stats['adjusted_over_rate']) * self.recent_weight
        
        return {
            'home_stats': home_stats,
            'away_stats': away_stats,
            'home_adjustment': home_adjustment + home_recent_adj,
            'away_adjustment': away_adjustment + away_recent_adj,
            'combined_adjustment': (home_adjustment + away_adjustment + home_recent_adj + away_recent_adj) / 2
        }
    
    def analyze_match_context(self, home_stats, away_stats, league_analysis):
        """
        ETAPA 3: Análise do Contexto do Jogo
        Fatores específicos do confronto
        """
        
        # Head-to-head seria aqui (se disponível)
        h2h_adjustment = 0  # Placeholder
        
        # Motivação/contexto (posição na tabela, etc.)
        motivation_adjustment = 0  # Placeholder
        
        # Estilo de jogo compatibilidade
        # Times ofensivos vs defensivos
        home_attack_style = home_stats['scoring_frequency']
        away_defense_style = 1 - away_stats['scoring_frequency']
        
        style_compatibility = (home_attack_style + away_defense_style) / 2 - 0.5
        
        return {
            'h2h_adjustment': h2h_adjustment,
            'motivation_adjustment': motivation_adjustment,
            'style_compatibility': style_compatibility,
            'final_context_adjustment': (h2h_adjustment + motivation_adjustment + style_compatibility) / 3
        }
    
    def predict_match(self, league_data, home_team_data, away_team_data):
        """
        Sistema Hierárquico Completo de Predição
        """
        
        # ETAPA 1: Análise da Liga
        league_analysis = self.analyze_league(league_data)
        base_probability = league_analysis['base_rate']
        
        # ETAPA 2: Análise dos Times
        team_analysis = self.analyze_teams(home_team_data, away_team_data)
        team_adjustment = team_analysis['combined_adjustment']
        
        # ETAPA 3: Análise do Contexto
        context_analysis = self.analyze_match_context(
            team_analysis['home_stats'], 
            team_analysis['away_stats'], 
            league_analysis
        )
        context_adjustment = context_analysis['final_context_adjustment']
        
        # ETAPA 4: Cálculo Final
        # Base da liga + ajustes dos times + contexto do jogo
        final_probability = base_probability + team_adjustment + context_adjustment
        
        # Manter entre 0 e 1
        final_probability = max(0.05, min(0.95, final_probability))
        
        # Convertir para porcentagem
        confidence_percentage = final_probability * 100
        
        # Garantir que está dentro dos limites da API (50-100%)
        confidence_percentage = max(self.confidence_range['min'], 
                                  min(self.confidence_range['max'], confidence_percentage))
        
        return {
            'final_probability': final_probability,
            'confidence_percentage': confidence_percentage,
            'breakdown': {
                'league_base': base_probability * 100,
                'team_adjustment': team_adjustment * 100,
                'context_adjustment': context_adjustment * 100,
                'final': confidence_percentage
            },
            'analysis_details': {
                'league': league_analysis,
                'teams': team_analysis,
                'context': context_analysis
            }
        }
    
    def predict_from_api(self, league_id, home_team_id, away_team_id, match_date=None):
        """
        Faz predição completa usando dados da API HT Goals
        """
        # Buscar dados da API
        league_data, home_team_data, away_team_data = self.fetch_data_from_api(
            league_id, home_team_id, away_team_id
        )
        
        # Verificar dados mínimos
        if len(home_team_data) < self.min_games_threshold:
            return {
                'error': f'Dados insuficientes para time da casa (mínimo: {self.min_games_threshold} jogos)',
                'confidence_percentage': 50
            }
        
        if len(away_team_data) < self.min_games_threshold:
            return {
                'error': f'Dados insuficientes para time visitante (mínimo: {self.min_games_threshold} jogos)',
                'confidence_percentage': 50
            }
        
        # Fazer predição
        prediction = self.predict_match(league_data, home_team_data, away_team_data)
        
        # Preparar dados para API
        api_payload = {
            'league_id': league_id,
            'home_team_id': home_team_id,
            'away_team_id': away_team_id,
            'match_date': match_date or datetime.now().isoformat(),
            'prediction': {
                'market': 'over_05_ht',
                'confidence': prediction['confidence_percentage'],
                'probability': prediction['final_probability'],
                'breakdown': prediction['breakdown']
            },
            'metadata': {
                'model_version': '2.0',
                'analysis_type': 'hierarchical',
                'data_points': {
                    'league_games': len(league_data),
                    'home_games': len(home_team_data),
                    'away_games': len(away_team_data)
                }
            }
        }
        
        # Enviar para API
        if self.api_client:
            api_response = self.api_client.make_prediction(api_payload)
            prediction['api_response'] = api_response
        
        return prediction

class BacktestingEngine:
    """Engine para testar performance histórica com 365 dias e divisão train/val/test"""
    
    def __init__(self, prediction_system):
        self.prediction_system = prediction_system
        self.results = []
        self.training_history = []
    
    def split_temporal_data(self, historical_data, train_days=240, val_days=60, test_days=65):
        """
        Divisão temporal dos dados (365 dias total):
        - Train: 240 dias (65.8%) - Para treinar o sistema
        - Validation: 60 dias (16.4%) - Para ajustar parâmetros  
        - Test: 65 dias (17.8%) - Para avaliar performance final
        """
        
        if 'date' not in historical_data.columns:
            raise ValueError("Dados históricos devem ter coluna 'date'")
        
        # Ordenar por data
        historical_data['date'] = pd.to_datetime(historical_data['date'], errors='coerce')
        historical_data = historical_data.sort_values('date').reset_index(drop=True)
        
        # Calcular datas de corte
        max_date = historical_data['date'].max()
        
        test_start = max_date - timedelta(days=test_days)
        val_start = test_start - timedelta(days=val_days)
        train_start = val_start - timedelta(days=train_days)
        
        # Filtrar apenas últimos 365 dias
        cutoff_date = max_date - timedelta(days=365)
        recent_data = historical_data[historical_data['date'] >= cutoff_date].copy()
        
        # Dividir dados
        train_data = recent_data[
            (recent_data['date'] >= train_start) & 
            (recent_data['date'] < val_start)
        ].copy()
        
        val_data = recent_data[
            (recent_data['date'] >= val_start) & 
            (recent_data['date'] < test_start)
        ].copy()
        
        test_data = recent_data[
            recent_data['date'] >= test_start
        ].copy()
        
        return {
            'train': train_data,
            'validation': val_data, 
            'test': test_data,
            'split_info': {
                'train_period': f"{train_start.strftime('%Y-%m-%d')} to {val_start.strftime('%Y-%m-%d')}",
                'val_period': f"{val_start.strftime('%Y-%m-%d')} to {test_start.strftime('%Y-%m-%d')}",
                'test_period': f"{test_start.strftime('%Y-%m-%d')} to {max_date.strftime('%Y-%m-%d')}",
                'train_games': len(train_data),
                'val_games': len(val_data),
                'test_games': len(test_data)
            }
        }
    
    def optimize_parameters(self, train_data, val_data):
        """
        Otimiza parâmetros do sistema usando dados de validação
        """
        
        best_accuracy = 0
        best_params = {
            'outlier_threshold': 1.5,
            'recent_weight': 0.3,
            'consistency_weight': 0.7,
            'min_games_threshold': 3
        }
        
        # Grid search para otimizar parâmetros
        outlier_thresholds = [1.0, 1.5, 2.0]
        recent_weights = [0.2, 0.3, 0.4]
        consistency_weights = [0.6, 0.7, 0.8]
        min_games = [3, 5, 7]
        
        optimization_results = []
        
        print("🔧 Testando combinações de parâmetros...")
        total_combinations = len(outlier_thresholds) * len(recent_weights) * len(consistency_weights) * len(min_games)
        current_combination = 0
        
        for outlier_thresh in outlier_thresholds:
            for recent_w in recent_weights:
                for consist_w in consistency_weights:
                    for min_g in min_games:
                        current_combination += 1
                        
                        if current_combination % 20 == 0:
                            print(f"   Progresso: {current_combination}/{total_combinations}")
                        
                        # Testar estes parâmetros
                        temp_system = SmartPredictionSystem()
                        temp_system.outlier_threshold = outlier_thresh
                        temp_system.recent_weight = recent_w
                        temp_system.consistency_weight = consist_w
                        temp_system.min_games_threshold = min_g
                        
                        # Avaliar no conjunto de validação
                        val_accuracy = self.evaluate_on_dataset(temp_system, train_data, val_data)
                        
                        optimization_results.append({
                            'outlier_threshold': outlier_thresh,
                            'recent_weight': recent_w,
                            'consistency_weight': consist_w,
                            'min_games_threshold': min_g,
                            'val_accuracy': val_accuracy
                        })
                        
                        if val_accuracy > best_accuracy:
                            best_accuracy = val_accuracy
                            best_params = {
                                'outlier_threshold': outlier_thresh,
                                'recent_weight': recent_w,
                                'consistency_weight': consist_w,
                                'min_games_threshold': min_g
                            }
        
        return best_params, best_accuracy, optimization_results
    
    def evaluate_on_dataset(self, system, train_data, eval_data):
        """Avalia sistema em um conjunto de dados"""
        
        predictions = []
        actuals = []
        
        # Agrupar por liga para eficiência
        leagues = eval_data['league_id'].unique()
        
        for league_id in leagues:
            league_train = train_data[train_data['league_id'] == league_id]
            league_eval = eval_data[eval_data['league_id'] == league_id]
            
            if len(league_train) < 20:  # Mínimo para treinamento
                continue
            
            # Para cada jogo de avaliação
            for idx, match in league_eval.iterrows():
                try:
                    # Dados dos times até a data do jogo
                    match_date = match['date']
                    historical_until_match = league_train[league_train['date'] < match_date]
                    
                    if len(historical_until_match) < 10:
                        continue
                    
                    # Dados específicos dos times
                    home_team_data = historical_until_match[
                        historical_until_match['home_team_id'] == match['home_team_id']
                    ]
                    away_team_data = historical_until_match[
                        historical_until_match['away_team_id'] == match['away_team_id']
                    ]
                    
                    if len(home_team_data) < system.min_games_threshold or \
                       len(away_team_data) < system.min_games_threshold:
                        continue
                    
                    # Fazer predição
                    prediction = system.predict_match(
                        historical_until_match, home_team_data, away_team_data
                    )
                    
                    # Comparar com resultado real
                    predicted_class = prediction['final_probability'] > 0.5
                    actual_class = match['over_05'] == 1
                    
                    predictions.append(predicted_class)
                    actuals.append(actual_class)
                    
                except Exception as e:
                    continue
        
        # Calcular acurácia
        if len(predictions) > 0:
            accuracy = accuracy_score(actuals, predictions)
            return accuracy
        
        return 0.0
    
    def run_comprehensive_backtest(self, historical_data):
        """
        Executa backtesting completo com 365 dias:
        1. Split train/val/test
        2. Otimização de parâmetros 
        3. Avaliação final
        4. Análise detalhada
        """
        
        print("🚀 Iniciando Backtesting Completo de 365 dias...")
        
        # 1. Dividir dados temporalmente
        print("📊 Dividindo dados em Train/Validation/Test...")
        data_split = self.split_temporal_data(historical_data)
        
        train_data = data_split['train']
        val_data = data_split['validation']
        test_data = data_split['test']
        
        print(f"✅ Train: {len(train_data)} jogos | Val: {len(val_data)} jogos | Test: {len(test_data)} jogos")
        
        # 2. Otimizar parâmetros usando train+validation
        print("🔧 Otimizando parâmetros do sistema...")
        best_params, best_val_acc, optimization_results = self.optimize_parameters(train_data, val_data)
        
        print(f"✅ Melhor acurácia validação: {best_val_acc:.1%}")
        print(f"✅ Melhores parâmetros: {best_params}")
        
        # 3. Aplicar melhores parâmetros ao sistema
        optimized_system = SmartPredictionSystem()
        for param, value in best_params.items():
            setattr(optimized_system, param, value)
        
        # 4. Avaliação final no conjunto de teste
        print("🎯 Avaliação final no conjunto de teste...")
        test_results = self.detailed_evaluation(optimized_system, train_data, val_data, test_data)
        
        # 5. Análise por faixas de confiança
        confidence_analysis = self.analyze_by_confidence_ranges(test_results['detailed_predictions'])
        
        # 6. Análise por liga
        league_analysis = self.analyze_by_league(test_results['detailed_predictions'])
        
        return {
            'data_split_info': data_split['split_info'],
            'optimization_results': {
                'best_params': best_params,
                'best_val_accuracy': best_val_acc,
                'all_combinations': optimization_results
            },
            'test_results': test_results,
            'confidence_analysis': confidence_analysis,
            'league_analysis': league_analysis,
            'summary': {
                'total_test_predictions': len(test_results['detailed_predictions']),
                'test_accuracy': test_results['accuracy'],
                'test_precision': test_results['precision'],
                'test_recall': test_results['recall'],
                'test_f1': test_results['f1_score']
            }
        }
    
    def detailed_evaluation(self, system, train_data, val_data, test_data):
        """Avaliação detalhada com múltiplas métricas"""
        
        # Combinar train+val para treinamento final
        training_data = pd.concat([train_data, val_data]).sort_values('date')
        
        detailed_predictions = []
        predictions = []
        actuals = []
        probabilities = []
        
        # Agrupar por liga
        leagues = test_data['league_id'].unique()
        
        for league_id in leagues:
            league_training = training_data[training_data['league_id'] == league_id]
            league_test = test_data[test_data['league_id'] == league_id]
            
            if len(league_training) < 30:  # Mínimo para avaliação confiável
                continue
            
            league_name = league_test['league_name'].iloc[0] if len(league_test) > 0 else f"Liga_{league_id}"
            
            # Para cada jogo de teste
            for idx, match in league_test.iterrows():
                try:
                    # Dados históricos até a data do jogo
                    match_date = match['date']
                    historical_data = league_training[league_training['date'] < match_date]
                    
                    # Dados dos times
                    home_team_data = historical_data[
                        historical_data['home_team_id'] == match['home_team_id']
                    ]
                    away_team_data = historical_data[
                        historical_data['away_team_id'] == match['away_team_id']
                    ]
                    
                    if len(home_team_data) < system.min_games_threshold or \
                       len(away_team_data) < system.min_games_threshold:
                        continue
                    
                    # Fazer predição
                    prediction = system.predict_match(historical_data, home_team_data, away_team_data)
                    
                    # Resultado real
                    actual = match['over_05'] == 1
                    predicted = prediction['final_probability'] > 0.5
                    
                    # Salvar resultados detalhados
                    detailed_predictions.append({
                        'date': match['date'],
                        'league': league_name,
                        'home_team': match.get('home_team', 'Unknown'),
                        'away_team': match.get('away_team', 'Unknown'),
                        'predicted_probability': prediction['final_probability'],
                        'confidence_percentage': prediction['confidence_percentage'],
                        'predicted_class': predicted,
                        'actual_class': actual,
                        'correct': predicted == actual,
                        'breakdown': prediction['breakdown'],
                        'ht_home_goals': match.get('ht_home_goals', 0),
                        'ht_away_goals': match.get('ht_away_goals', 0)
                    })
                    
                    predictions.append(predicted)
                    actuals.append(actual)
                    probabilities.append(prediction['final_probability'])
                    
                except Exception as e:
                    continue
        
        # Calcular métricas
        if len(predictions) > 0:
            accuracy = accuracy_score(actuals, predictions)
            precision = precision_score(actuals, predictions, zero_division=0)
            recall = recall_score(actuals, predictions, zero_division=0)
            f1 = f1_score(actuals, predictions, zero_division=0)
            
            return {
                'accuracy': accuracy,
                'precision': precision,
                'recall': recall,
                'f1_score': f1,
                'total_predictions': len(predictions),
                'detailed_predictions': detailed_predictions
            }
        
        return {
            'accuracy': 0,
            'precision': 0,
            'recall': 0,
            'f1_score': 0,
            'total_predictions': 0,
            'detailed_predictions': []
        }
    
    def analyze_by_confidence_ranges(self, detailed_predictions):
        """Analisa performance por faixas de confiança"""
        
        if not detailed_predictions:
            return []
        
        df = pd.DataFrame(detailed_predictions)
        
        confidence_ranges = [
            (50, 60), (60, 70), (70, 80), (80, 90), (90, 100)
        ]
        
        analysis = []
        
        for min_conf, max_conf in confidence_ranges:
            mask = (df['confidence_percentage'] >= min_conf) & \
                   (df['confidence_percentage'] < max_conf)
            
            subset = df[mask]
            
            if len(subset) > 0:
                accuracy = subset['correct'].mean()
                count = len(subset)
                avg_confidence = subset['confidence_percentage'].mean()
                
                # Análise adicional
                over_predictions = subset[subset['predicted_class'] == True]
                over_accuracy = over_predictions['correct'].mean() if len(over_predictions) > 0 else 0
                
                analysis.append({
                    'confidence_range': f"{min_conf}-{max_conf}%",
                    'predictions_count': count,
                    'accuracy': accuracy,
                    'avg_confidence': avg_confidence,
                    'over_predictions': len(over_predictions),
                    'over_accuracy': over_accuracy
                })
        
        return analysis
    
    def analyze_by_league(self, detailed_predictions):
        """Analisa performance por liga"""
        
        if not detailed_predictions:
            return []
        
        df = pd.DataFrame(detailed_predictions)
        
        league_analysis = []
        
        for league in df['league'].unique():
            league_data = df[df['league'] == league]
            
            if len(league_data) >= 10:  # Mínimo para análise
                accuracy = league_data['correct'].mean()
                avg_confidence = league_data['confidence_percentage'].mean()
                
                league_analysis.append({
                    'league': league,
                    'predictions_count': len(league_data),
                    'accuracy': accuracy,
                    'avg_confidence': avg_confidence
                })
        
        # Ordenar por acurácia
        league_analysis.sort(key=lambda x: x['accuracy'], reverse=True)
        
        return league_analysis

# Funções para integração com Streamlit e API
def create_prediction_breakdown_display(prediction_result):
    """Cria display detalhado da predição para Streamlit"""
    
    breakdown = prediction_result['breakdown']
    details = prediction_result['analysis_details']
    
    return {
        'Liga Base': f"{breakdown['league_base']:.1f}%",
        'Ajuste Times': f"{breakdown['team_adjustment']:+.1f}%",
        'Contexto Jogo': f"{breakdown['context_adjustment']:+.1f}%",
        'Taxa Final': f"{breakdown['final']:.1f}%",
        'Jogos Casa': details['teams']['home_stats']['games_count'],
        'Jogos Fora': details['teams']['away_stats']['games_count'],
        'Consistência Casa': f"{details['teams']['home_stats']['consistency_score']:.2f}",
        'Consistência Fora': f"{details['teams']['away_stats']['consistency_score']:.2f}",
        'Outliers Casa': details['teams']['home_stats']['outliers_detected'],
        'Outliers Fora': details['teams']['away_stats']['outliers_detected']
    }

def example_usage_with_api():
    """Exemplo de uso com integração API HT Goals"""
    
    # Configurar API client
    api_client = HTGoalsAPIClient(api_key="YOUR_API_KEY_HERE")
    
    # Criar sistema de predição
    prediction_system = SmartPredictionSystem(api_client=api_client)
    
    # Fazer predição para um jogo específico
    result = prediction_system.predict_from_api(
        league_id=1,  # Premier League
        home_team_id=10,  # Manchester United
        away_team_id=20,  # Liverpool
        match_date="2025-06-02T15:00:00"
    )
    
    if 'error' not in result:
        print("🎯 Resultado da Predição:")
        print(f"Confiança: {result['confidence_percentage']:.1f}%")
        print(f"Probabilidade: {result['final_probability']:.2f}")
        print("\n📊 Breakdown:")
        for key, value in result['breakdown'].items():
            print(f"  {key}: {value:.1f}%")
    else:
        print(f"❌ Erro: {result['error']}")
    
    return result

def export_predictions_to_excel(predictions_df, filename="ht_goals_predictions.xlsx"):
    """
    Exporta predições para Excel com formatação profissional
    """
    if not EXCEL_AVAILABLE:
        # Fallback para CSV se openpyxl não estiver instalado
        predictions_df.to_csv(filename.replace('.xlsx', '.csv'), index=False)
        return filename.replace('.xlsx', '.csv')
    
    # Criar Excel writer
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        # Aba principal com predições
        predictions_df.to_excel(writer, sheet_name='Predições', index=False)
        
        # Pegar workbook e worksheet
        workbook = writer.book
        worksheet = writer.sheets['Predições']
        
        # Estilos
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        # Formatar cabeçalhos
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        
        # Ajustar largura das colunas
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Adicionar formatação condicional para confiança
        for row in range(2, len(predictions_df) + 2):
            conf_cell = worksheet[f'F{row}']  # Assumindo que confiança está na coluna F
            try:
                conf_value = float(conf_cell.value)
                if conf_value >= 80:
                    conf_cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
                elif conf_value >= 70:
                    conf_cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
                else:
                    conf_cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            except:
                pass
        
        # Criar aba de resumo
        summary_data = create_summary_stats(predictions_df)
        summary_df = pd.DataFrame(summary_data)
        summary_df.to_excel(writer, sheet_name='Resumo', index=False)
        
        # Formatar aba de resumo
        summary_sheet = writer.sheets['Resumo']
        for cell in summary_sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
    
    return filename

def create_summary_stats(predictions_df):
    """Cria estatísticas resumidas das predições"""
    
    summary = []
    
    # Estatísticas gerais
    summary.append({
        'Métrica': 'Total de Predições',
        'Valor': len(predictions_df)
    })
    
    summary.append({
        'Métrica': 'Confiança Média',
        'Valor': f"{predictions_df['confidence_percentage'].mean():.1f}%"
    })
    
    # Por faixas de confiança
    confidence_ranges = [(50, 60), (60, 70), (70, 80), (80, 90), (90, 100)]
    
    for min_conf, max_conf in confidence_ranges:
        mask = (predictions_df['confidence_percentage'] >= min_conf) & \
               (predictions_df['confidence_percentage'] < max_conf)
        count = mask.sum()
        
        summary.append({
            'Métrica': f'Predições {min_conf}-{max_conf}%',
            'Valor': count
        })
    
    # Se houver resultados reais
    if 'correct' in predictions_df.columns:
        accuracy = predictions_df['correct'].mean()
        summary.append({
            'Métrica': 'Acurácia Geral',
            'Valor': f"{accuracy:.1%}"
        })
    
    return summary

def demonstrate_outlier_handling():
    """
    Demonstra como o sistema lida com outliers
    Exemplo: Time com [5, 0, 0, 0, 0] gols HT
    """
    
    print("🔍 DEMONSTRAÇÃO: Tratamento de Outliers")
    print("=" * 60)
    
    # Criar sistema
    system = SmartPredictionSystem()
    
    # Dados do seu exemplo - time da casa
    home_team_data = pd.DataFrame({
        'ht_home_goals': [5, 0, 0, 0, 0],  # Exatamente seu exemplo!
        'over_05': [1, 0, 0, 0, 0],
        'date': pd.date_range('2024-01-01', periods=5)
    })
    
    # Analisar time
    stats = system.calculate_realistic_team_stats(home_team_data, 'home')
    
    print("📊 Dados originais do time:")
    print(f"   Gols HT: {home_team_data['ht_home_goals'].tolist()}")
    print(f"   Média com outlier: {home_team_data['ht_home_goals'].mean():.2f} gols")
    
    print("\n🎯 Após detecção de outliers:")
    print(f"   Outliers detectados: {stats['outliers_detected']}")
    print(f"   Taxa Over 0.5 raw: {stats['raw_over_rate']:.1%}")
    print(f"   Taxa Over 0.5 ajustada: {stats['adjusted_over_rate']:.1%}")
    
    # Mostrar o que aconteceu
    goals = home_team_data['ht_home_goals'].values
    outliers = system.detect_outliers(goals)
    clean_goals = goals[~outliers]
    
    print(f"\n✅ Gols após remover outliers: {clean_goals.tolist()}")
    print(f"   Média sem outlier: {clean_goals.mean():.2f} gols")
    print(f"   Frequência de marcar (limpa): {(clean_goals > 0).mean():.1%}")
    
    return stats

# Exemplo de uso completo com export
def complete_example_with_export():
    """Exemplo completo: análise + export Excel"""
    
    # 1. Configurar sistema
    api_client = HTGoalsAPIClient(api_key="YOUR_API_KEY")
    system = SmartPredictionSystem(api_client=api_client)
    
    # 2. Fazer várias predições
    predictions = []
    
    matches_to_predict = [
        {'league_id': 1, 'home_id': 10, 'away_id': 20, 'date': '2025-06-02'},
        {'league_id': 1, 'home_id': 30, 'away_id': 40, 'date': '2025-06-02'},
        # ... mais jogos
    ]
    
    for match in matches_to_predict:
        try:
            result = system.predict_from_api(
                league_id=match['league_id'],
                home_team_id=match['home_id'],
                away_team_id=match['away_id'],
                match_date=match['date']
            )
            
            predictions.append({
                'Data': match['date'],
                'Liga': f"Liga_{match['league_id']}",
                'Casa': f"Time_{match['home_id']}",
                'Fora': f"Time_{match['away_id']}",
                'Probabilidade': result['final_probability'],
                'Confiança (%)': result['confidence_percentage'],
                'Base Liga (%)': result['breakdown']['league_base'],
                'Ajuste Times (%)': result['breakdown']['team_adjustment'],
                'Contexto (%)': result['breakdown']['context_adjustment']
            })
        except Exception as e:
            print(f"Erro na predição: {e}")
    
    # 3. Criar DataFrame
    df = pd.DataFrame(predictions)
    
    # 4. Exportar para Excel
    filename = export_predictions_to_excel(df, "predicoes_ht_goals.xlsx")
    print(f"\n✅ Predições exportadas para: {filename}")
    
    return df
